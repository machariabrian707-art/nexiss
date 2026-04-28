"""Tests for the xlsx export service.

Fixed: test_export_empty_list_raises now correctly tests that an empty list
produces a valid header-only xlsx (no IndexError).
"""
from __future__ import annotations

import io
import uuid
from datetime import datetime, timezone
from unittest.mock import MagicMock

from nexiss.db.models.document import DocumentStatus


def _make_doc(
    *,
    file_name: str = "invoice.pdf",
    confirmed_type: str = "business_financial",
    extracted_data: dict | None = None,
) -> MagicMock:
    doc = MagicMock()
    doc.id = uuid.uuid4()
    doc.file_name = file_name
    doc.status = DocumentStatus.completed
    doc.declared_type = confirmed_type
    doc.confirmed_type = confirmed_type
    doc.document_subtype = "invoice"
    doc.page_count = 2
    doc.created_at = datetime(2026, 1, 15, tzinfo=timezone.utc)
    doc.extracted_data = extracted_data or {
        "vendor_name": "Doshi Traders",
        "total_amount": "5000",
        "currency": "KES",
    }
    return doc


def test_export_produces_xlsx_bytes():
    from nexiss.services.export.xlsx_export import export_documents_to_xlsx
    docs = [_make_doc(), _make_doc(file_name="receipt.pdf")]
    result = export_documents_to_xlsx(docs)
    assert isinstance(result, bytes)
    assert result[:2] == b"PK"  # xlsx is a zip


def test_export_xlsx_is_parseable():
    import openpyxl
    from nexiss.services.export.xlsx_export import export_documents_to_xlsx
    docs = [_make_doc(extracted_data={"vendor_name": "Test Corp", "amount": "100"})]
    xlsx_bytes = export_documents_to_xlsx(docs)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "file_name" in headers
    assert "confirmed_type" in headers
    assert ws.max_row == 2


def test_export_includes_extracted_fields_as_columns():
    import openpyxl
    from nexiss.services.export.xlsx_export import export_documents_to_xlsx
    docs = [_make_doc(extracted_data={"vendor_name": "Doshi", "total_amount": "9999"})]
    xlsx_bytes = export_documents_to_xlsx(docs)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "vendor_name" in headers
    assert "total_amount" in headers


def test_export_medical_records_sorted_by_patient():
    import openpyxl
    from nexiss.services.export.xlsx_export import export_documents_to_xlsx
    doc_b = _make_doc(
        file_name="bob.pdf",
        confirmed_type="medical_healthcare",
        extracted_data={"patient_name": "Bob"},
    )
    doc_a = _make_doc(
        file_name="alice.pdf",
        confirmed_type="medical_healthcare",
        extracted_data={"patient_name": "Alice"},
    )
    xlsx_bytes = export_documents_to_xlsx([doc_b, doc_a])
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    fn_col = headers.index("file_name") + 1
    assert ws.cell(2, fn_col).value == "alice.pdf"
    assert ws.cell(3, fn_col).value == "bob.pdf"


def test_export_empty_list_returns_header_only_xlsx():
    """Empty document list must not raise — returns a valid xlsx with header row only."""
    import openpyxl
    from nexiss.services.export.xlsx_export import export_documents_to_xlsx
    result = export_documents_to_xlsx([])
    assert isinstance(result, bytes)
    assert result[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    # Only header row, no data rows
    assert ws.max_row == 1
