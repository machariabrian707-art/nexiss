"""Export service: converts processed documents to Excel or CSV.

Supports:
  - Generic export (all fields in extracted_data as columns).
  - Medical-aware export (groups rows by patient, sorts by date).
  - Financial-aware export (invoices, receipts with amount columns).
  - The calling API endpoint decides which mode based on confirmed_type.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    _HAS_OPENPYXL = False

from nexiss.db.models.document import Document


_MEDICAL_CATEGORIES = {"medical_healthcare"}
_FINANCIAL_CATEGORIES = {
    "business_financial",
    "logistics_supply_chain",
}


def _flatten(data: dict | None, prefix: str = "") -> dict[str, Any]:
    """Flatten nested dicts into dot-separated column names."""
    result: dict[str, Any] = {}
    if not data:
        return result
    for key, val in data.items():
        full_key = f"{prefix}.{key}" if prefix else key
        if isinstance(val, dict):
            result.update(_flatten(val, full_key))
        elif isinstance(val, list):
            result[full_key] = ", ".join(str(v) for v in val)
        else:
            result[full_key] = val
    return result


def _medical_sort_key(row: dict) -> tuple:
    patient = str(row.get("patient_name") or row.get("patient") or "").lower()
    date_str = str(row.get("date") or row.get("visit_date") or row.get("created_at") or "")
    return (patient, date_str)


def export_documents_to_xlsx(documents: list[Document]) -> bytes:
    """Returns raw bytes of an .xlsx file."""
    if not _HAS_OPENPYXL:
        raise RuntimeError(
            "openpyxl is not installed. Add 'openpyxl>=3.1' to project dependencies."
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Documents"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    rows: list[dict[str, Any]] = []
    all_keys: list[str] = []
    seen_keys: set[str] = set()

    for doc in documents:
        base = {
            "id": str(doc.id),
            "file_name": doc.file_name,
            "status": doc.status.value if doc.status else "",
            "declared_type": doc.declared_type or "",
            "confirmed_type": doc.confirmed_type or "",
            "document_subtype": doc.document_subtype or "",
            "page_count": doc.page_count,
            "created_at": doc.created_at.isoformat() if doc.created_at else "",
        }
        flattened = _flatten(doc.extracted_data)
        merged = {**base, **flattened}
        rows.append(merged)
        for k in flattened:
            if k not in seen_keys:
                all_keys.append(k)
                seen_keys.add(k)

    # Determine sort strategy based on first doc's confirmed_type
    first_type = documents[0].confirmed_type if documents else ""
    if first_type in _MEDICAL_CATEGORIES:
        rows.sort(key=_medical_sort_key)

    base_cols = ["id", "file_name", "status", "declared_type", "confirmed_type",
                 "document_subtype", "page_count", "created_at"]
    all_columns = base_cols + [k for k in all_keys if k not in base_cols]

    # Write header
    for col_idx, col_name in enumerate(all_columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    # Write rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(all_columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
